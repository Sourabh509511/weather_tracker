import requests
import openpyxl
import os

def getdata(city,unit):
    '''this function gets data from the api'''
    if unit == "C":
        full_url=api_url+city+"&units=metric"
    elif unit=="F":
        full_url=api_url+city+"&units=imperial"
    json_data=requests.get(full_url).json()
    try:
        temperature=json_data["main"]["temp"]
        humidity=json_data["main"]["humidity"]
    except Exception as e:
        error_data=json_data["message"]
        print(error_data)

    return temperature,humidity

def write(data,row=None):
    '''this function updates weather data in sheet'''
    file_name=os.path.join(os.getcwd(),"Fox trading\\weather.xlsx")
    open_file=openpyxl.load_workbook(file_name)
    if len(data)>2:
        write_obj=open_file[str('Weather')]
        last_row = write_obj.max_row
        for i in range(1,6):
            write_data = write_obj.cell(row= last_row+1 , column = i)
            write_data.value = data[i-1]
            open_file.save(file_name)
    else:
        write_obj=open_file[str('Weather')]
        data_list=[data[0],data[1]]
        for i in range(2,4):
            write_data = write_obj.cell(row=row  , column = i)
            write_data.value = data_list[i-2]
            open_file.save(file_name)

api_url='https://api.openweathermap.org/data/2.5/weather?appid=7d7da215aa40a89fd48f5d0335687463&q='

a=int(input("Enter 1 to add cities:\n"))

if a==1:
    no_of_cities=int(input("Enter no of cities you want to track the weather\n").strip())

    for _ in range(no_of_cities):

        city=input("Enter city name\n").capitalize().strip()
        unit=input("Enter unit for temperature:\n for celsius (C)\n for farenheit (F)\n").capitalize().strip()

        if unit != "C" and unit!="F":
            print("Incorrect choice")
            break

        temperature,humidity=getdata(city,unit)

        update_choice=int(input(f"Do you want to get the update of {city} temperature every second\nEnter 1 for yes\n Enter 0 for no.\n").strip())
        file_name=os.path.join(os.getcwd(),"Fox trading\\weather.xlsx")
        open_file=openpyxl.load_workbook(file_name)
        token_obj=open_file[str('city tokens')]
        city1=[]
        for row in token_obj.rows:
            if row[0].value == f"{city}":
                for cell in row:
                    city1.append(cell.value)

        if len(city1)>2:
            print("There are more than one city with same name choose your city\n")
            for i in range(0,len(city1),2):
                print(f'city is:{city1[i]} and code is {city1[i+1]}.Enter {i} for  this')
            option=int(input().strip())
            code=city1[option+1]
        else:
            code=city1[1]
        data_list=[code,temperature,humidity,unit,update_choice]
        write(data_list)
        # write_obj=open_file[str('Weather')]
        # last_row = write_obj.max_row
        # for i in range(1,6):
        #     data = write_obj.cell(row= last_row+1 , column = i)
        #     data.value = data_list[i-1]
        #     open_file.save(file_name)

else:
    file_name=os.path.join(os.getcwd(),"Fox trading\\weather.xlsx")
    open_file=openpyxl.load_workbook(file_name)
    write_obj=open_file[str('Weather')]
    city=[]
    unit=[]
    for row in write_obj.rows:
        if row[4].value == 1:
            row_no=write_obj._current_row-1
            unit.append(row[3].value)
            token_value=row[0].value
            token_obj=open_file[str('city tokens')]
            for row in token_obj.rows:
                # print(f"{token_value} {row[1].value}")
                if row[1].value == f"{token_value}":
                    city.append(row[0].value)

    while(1):
        for i in range(len(city)):
            temp,humid=getdata(city[i],unit[i])
            data=[temp,humid]
            write(data,row_no)


